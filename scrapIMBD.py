from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "TOP RATED MOVIES"
# sheet.cell(row = 2, column = 2)
# print(excel.sheetnames)
sheet.append(["Movie Rank", "Movie Name", "Movie Year", "IMDB Rating", "Movie Link"])


try:
    url = requests.get("https://www.imdb.com/chart/top/")
    url.raise_for_status()  # It will throw an error if url is not right
    html_text = url.content

    soup = BeautifulSoup(html_text, 'html5lib')
    movies = soup.find('tbody', class_='lister-list').find_all('tr')
    # print(movies)
    # movies.find_all('tr')
    for movie in movies:
        # movie_name = movie.find('td', class_='titleColumn')
        # print(movie_name.text)
        movie_name = movie.find('td', class_='titleColumn').a.text

        rank = movie.find('td', class_='titleColumn').text.split('.')[0]

        year = movie.find('span', class_='secondaryInfo').text

        rating = movie.find('td', class_='ratingColumn imdbRating').text

        link = movie.a.img['src']

        sheet.append([rank, movie_name, year, rating, link])

        # print(f'Rank of the movie is: {rank.strip()}\n'
        #       f'Name of the movie is: {movie_name}\n'
        #       f'Year of the movie is: {year}\n'
        #       f'Rating of the movie is: {ratings.strip()}\n')
        # print('\n')

except Exception as e:
    print(e)

excel.save('IMBD TOP 250 MOVIES.xlsx')
print("File Saved!")